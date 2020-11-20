import openpyxl #Add the library we need to process excel files
from datetime import date
from dateutil import parser as dateParser
from utils import to_excel_date
from utils import is_date

#Create a variable wb to hold an instance of the excel file to process
modelTemplateWorkbook = openpyxl.load_workbook('ModelTemplate.xlsx')
#Create a variable with all the worksheets from this workbook
sheets = modelTemplateWorkbook._sheets
for sheet in sheets:
    print("Sheet: \t" + sheet.title)
    for i in range(2, 734):
        if is_date(str(sheet.cell(row=i, column=1).value)):
            dateFromExcel = dateParser.parse(str(sheet.cell(row=i, column=1).value))
            print("ToExcelFormat:\t" + to_excel_date(dateFromExcel))
            print("IsoFormat:\t"  +  dateFromExcel.isoformat())
            print("RegFormat:\t"  +  str(sheet.cell(row=i, column=1).value))
    # b738 = sheet["B738"].value
    # print("B738: \t" + str(b738))
    print('\n')
